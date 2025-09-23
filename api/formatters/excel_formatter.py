#!/usr/bin/env python3
"""
Excel-specific formatting utilities
Separated from transaction_formatter to avoid loading Excel dependencies in non-Excel Lambda functions
"""

from typing import Dict, List, Any
from datetime import datetime
from io import BytesIO
from .transaction_formatter import format_transaction_for_display


def get_statement_filename(statement_metadata: Dict[str, Any], job_id: str, extension: str = 'xlsx') -> str:
    """
    Generate appropriate filename for Excel export

    Args:
        statement_metadata: Statement metadata from processing
        job_id: Job ID for fallback
        extension: File extension (xlsx)

    Returns:
        Filename for download
    """
    try:
        import re

        bank_name = statement_metadata.get('bank_name', 'Bank')
        account_number = statement_metadata.get('account_number', '')
        statement_period = statement_metadata.get('statement_period', {})

        # Clean bank name for filename
        bank_clean = re.sub(r'[^\w\s-]', '', bank_name).strip()
        bank_clean = re.sub(r'\s+', '_', bank_clean)

        # Get date range
        from_date = statement_period.get('from_date', '')
        to_date = statement_period.get('to_date', '')

        if from_date and to_date:
            # Convert dates to YYYY-MM-DD format for filename
            try:
                from_formatted = datetime.strptime(from_date, "%d/%m/%Y").strftime("%Y-%m-%d")
                to_formatted = datetime.strptime(to_date, "%d/%m/%Y").strftime("%Y-%m-%d")
                date_range = f"{from_formatted}_to_{to_formatted}"
            except:
                date_range = f"{from_date.replace('/', '-')}_to_{to_date.replace('/', '-')}"
        else:
            date_range = datetime.now().strftime("%Y-%m-%d")

        # Build filename
        if account_number:
            acc_suffix = account_number[-4:] if len(account_number) >= 4 else account_number
            filename = f"{bank_clean}_Statement_AC_{acc_suffix}_{date_range}.{extension}"
        else:
            filename = f"{bank_clean}_Statement_{date_range}.{extension}"

        return filename

    except Exception:
        # Fallback to simple filename
        return f"bank_statement_{job_id}.{extension}"


def create_excel_workbook(transactions: List[Dict[str, Any]]) -> BytesIO:
    """
    Generate Excel content with formatting

    Args:
        transactions: List of raw transaction data

    Returns:
        BytesIO buffer containing Excel file
    """
    # Import openpyxl at runtime to avoid module-level import issues
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError(f"openpyxl is not available: {e}")

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    # Define colors and alignment
    red_font = Font(color="FF0000")  # Red for debits
    green_font = Font(color="008000")  # Green for credits
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")

    # Add headers
    headers = ['Txn Date', 'Value Date', 'Description', 'Debit', 'Credit', 'Balance']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        # Right-align Debit, Credit, and Balance headers (columns 4, 5, 6)
        if col_num in [4, 5, 6]:
            cell.alignment = right_align
        else:
            cell.alignment = center_align

    # Add data rows
    row_num = 2
    for transaction in transactions:
        formatted = format_transaction_for_display(transaction)

        # Txn Date (DD-MM-YYYY format)
        ws.cell(row=row_num, column=1, value=formatted['txn_date'])

        # Value Date (DD-MM-YYYY format)
        ws.cell(row=row_num, column=2, value=formatted['value_date'])

        # Description
        ws.cell(row=row_num, column=3, value=formatted['description'])

        # Debit - with red color if has value, no formatting/rounding, right-aligned
        debit_value = formatted['debit']
        debit_cell = ws.cell(row=row_num, column=4, value=debit_value if debit_value else '')
        debit_cell.alignment = right_align
        if debit_value:
            debit_cell.font = red_font

        # Credit - with green color if has value, no formatting/rounding, right-aligned
        credit_value = formatted['credit']
        credit_cell = ws.cell(row=row_num, column=5, value=credit_value if credit_value else '')
        credit_cell.alignment = right_align
        if credit_value:
            credit_cell.font = green_font

        # Balance - no formatting/rounding, right-aligned
        balance_cell = ws.cell(row=row_num, column=6, value=formatted['balance'])
        balance_cell.alignment = right_align

        row_num += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Set minimum width and add some padding
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO buffer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer